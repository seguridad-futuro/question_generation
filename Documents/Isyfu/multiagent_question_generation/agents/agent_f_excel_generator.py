"""Agente F: Excel Generator - Genera Excels de examenes por tema

Arquitectura usando LangGraph:
- StateGraph para flujo de nodos
- Agrupacion de preguntas por tema
- Generacion de Excel con formato profesional (openpyxl)
- Un archivo .xlsx por cada tema
"""

import logging
import re
from pathlib import Path
from typing import List, Dict, Any, Optional, TypedDict
from datetime import datetime
from enum import Enum
from pydantic import BaseModel, Field

# LangGraph Imports
from langgraph.graph import StateGraph, END

# openpyxl para Excel
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter

# Imports del proyecto
from models.question import Question


# ==========================================
# CONFIGURACION DE LOGGING
# ==========================================

def setup_logger(name: str = "AgentF", level: int = logging.INFO) -> logging.Logger:
    logger = logging.getLogger(name)
    logger.setLevel(level)
    if not logger.handlers:
        handler = logging.StreamHandler()
        formatter = logging.Formatter(
            '\n%(asctime)s | %(name)s | %(levelname)s\n%(message)s',
            datefmt='%H:%M:%S'
        )
        handler.setFormatter(formatter)
        logger.addHandler(handler)
    return logger


logger = setup_logger("AgentF")


def log_separator(title: str = "", char: str = "=", width: int = 70):
    if title:
        padding = (width - len(title) - 2) // 2
        print(f"\n{char * padding} {title} {char * padding}")
    else:
        print(f"\n{char * width}")


def _strip_option_prefix(text: Optional[str]) -> str:
    """Elimina etiquetas tipo 'A)', 'B.' o checkmarks al inicio de una opcion."""
    if not text:
        return ""
    cleaned = text.strip()
    cleaned = re.sub(r"^[^A-Za-z0-9]*\s*", "", cleaned)
    label_re = re.compile(r"^\s*(?:\(?[A-Da-d]|[1-4]\)?)[.)\-:]+\s*")
    while True:
        updated = label_re.sub("", cleaned, count=1)
        if updated == cleaned:
            break
        cleaned = updated.strip()
    return cleaned


# ==========================================
# 1. MODELOS DE DATOS
# ==========================================

class ExcelFormatEnum(str, Enum):
    EXAM = "exam"
    WITH_SOLUTIONS = "with_solutions"
    STUDY_GUIDE = "study_guide"


class ExcelMetadata(BaseModel):
    topic: int
    total_questions: int
    format: ExcelFormatEnum
    generated_at: str
    file_name: str
    file_path: str


# ==========================================
# 2. ESTADO DEL AGENTE
# ==========================================

class AgentFState(TypedDict):
    questions: List[Question]
    output_dir: Path
    excel_format: ExcelFormatEnum
    topic_filter: Optional[int]

    questions_by_topic: Dict[int, List[Question]]
    excel_metadata: List[ExcelMetadata]

    decisions_log: List[Dict[str, Any]]
    node_count: int


# ==========================================
# 3. UTILIDADES
# ==========================================

def add_decision_log(state: AgentFState, node: str, decision: str, details: Dict = None) -> List[Dict]:
    log_entry = {
        "timestamp": datetime.now().isoformat(),
        "node": node,
        "decision": decision,
        "details": details or {}
    }
    decisions_log = state.get("decisions_log", [])
    decisions_log.append(log_entry)
    return decisions_log


# ==========================================
# 4. NODOS DEL GRAFO
# ==========================================

def group_by_topic_node(state: AgentFState) -> dict:
    """NODO 1: Agrupa preguntas por tema."""
    log_separator("NODO: GROUP BY TOPIC", "=")

    questions = state["questions"]
    topic_filter = state.get("topic_filter")

    print(f"\nAgrupando {len(questions)} preguntas (filtro tema: {topic_filter or 'todos'})")

    questions_by_topic: Dict[int, List[Question]] = {}

    for question in questions:
        topic = question.topic
        if topic_filter is not None and topic != topic_filter:
            continue
        if topic not in questions_by_topic:
            questions_by_topic[topic] = []
        questions_by_topic[topic].append(question)

    print(f"Temas encontrados: {len(questions_by_topic)}")
    for topic, qs in sorted(questions_by_topic.items()):
        print(f"  Tema {topic}: {len(qs)} preguntas")

    return {
        "questions_by_topic": questions_by_topic,
        "decisions_log": add_decision_log(state, "group_by_topic", "completed", {
            "unique_topics": len(questions_by_topic),
            "total_questions": sum(len(qs) for qs in questions_by_topic.values())
        }),
        "node_count": state.get("node_count", 0) + 1
    }


def generate_excels_node(state: AgentFState) -> dict:
    """NODO 2: Genera un Excel por cada tema."""
    log_separator("NODO: GENERATE EXCELS", "=")

    questions_by_topic = state["questions_by_topic"]
    output_dir = Path(state["output_dir"])
    excel_format = state["excel_format"]

    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\nGenerando Excels por tema:")
    print(f"  Temas: {len(questions_by_topic)}")
    print(f"  Formato: {excel_format.value}")
    print(f"  Directorio: {output_dir}")

    excel_metadata_list = []

    for topic, questions in sorted(questions_by_topic.items()):
        print(f"\n{'_'*50}")
        print(f"Generando Excel para Tema {topic} ({len(questions)} preguntas)...")

        try:
            excel_path = _generate_topic_excel(
                topic=topic,
                questions=questions,
                output_dir=output_dir,
                excel_format=excel_format
            )

            metadata = ExcelMetadata(
                topic=topic,
                total_questions=len(questions),
                format=excel_format,
                generated_at=datetime.now().isoformat(),
                file_name=excel_path.name,
                file_path=str(excel_path)
            )
            excel_metadata_list.append(metadata)
            print(f"Excel generado: {excel_path.name}")

        except Exception as e:
            logger.error(f"Error generando Excel para tema {topic}: {e}")
            import traceback
            traceback.print_exc()

    print(f"\n{'='*50}")
    print(f"Excels generados: {len(excel_metadata_list)}")

    return {
        "excel_metadata": excel_metadata_list,
        "decisions_log": add_decision_log(state, "generate_excels", "completed", {
            "excels_generated": len(excel_metadata_list)
        }),
        "node_count": state.get("node_count", 0) + 1
    }


# ==========================================
# 5. GENERACION DE EXCEL POR TEMA
# ==========================================

# Estilos reutilizables
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CORRECT_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
REVIEW_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _get_headers(excel_format: ExcelFormatEnum) -> List[str]:
    """Devuelve las cabeceras segun el formato."""
    base = ["#", "Pregunta"]

    if excel_format == ExcelFormatEnum.EXAM:
        return base + ["Opcion A", "Opcion B", "Opcion C", "Opcion D"]

    if excel_format == ExcelFormatEnum.WITH_SOLUTIONS:
        return base + [
            "Opcion A", "Opcion B", "Opcion C", "Opcion D",
            "Respuesta Correcta", "Tip", "Articulo"
        ]

    # STUDY_GUIDE
    return base + [
        "Opcion A", "Opcion B", "Opcion C", "Opcion D",
        "Respuesta Correcta", "Tip", "Articulo",
        "Dificultad", "Razon Dificultad",
        "Revision Manual", "Comentario Evaluacion",
        "Modelo LLM", "Faithfulness", "Relevancy",
        "Tiempo Gen (s)", "Tiempo Rev (s)"
    ]


def _get_column_widths(excel_format: ExcelFormatEnum) -> List[int]:
    """Anchos de columna segun formato."""
    base = [5, 50]

    if excel_format == ExcelFormatEnum.EXAM:
        return base + [30, 30, 30, 30]

    if excel_format == ExcelFormatEnum.WITH_SOLUTIONS:
        return base + [30, 30, 30, 30, 20, 50, 30]

    return base + [30, 30, 30, 30, 20, 50, 30, 12, 30, 15, 40, 15, 12, 12, 12, 12]


def _generate_topic_excel(
    topic: int,
    questions: List[Question],
    output_dir: Path,
    excel_format: ExcelFormatEnum
) -> Path:
    """Genera un archivo Excel para un tema especifico."""

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"Tema_{topic}_{excel_format.value}_{timestamp}.xlsx"
    excel_path = output_dir / file_name

    wb = Workbook()
    ws = wb.active
    ws.title = f"Tema {topic}"

    # Cabeceras
    headers = _get_headers(excel_format)
    col_widths = _get_column_widths(excel_format)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Anchos de columna
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Fijar fila de cabecera
    ws.freeze_panes = "A2"

    # Autofiltro
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Datos
    for row_idx, q in enumerate(questions, 2):
        correct_letter = ["A", "B", "C", "D"][q.solution - 1]
        correct_text = f"{correct_letter}) {_strip_option_prefix(q.get_correct_answer())}"

        row_data = _build_row(row_idx - 1, q, correct_text, excel_format)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGNMENT
            cell.border = THIN_BORDER

        # Colorear fila si necesita revision manual
        if getattr(q, "needs_manual_review", False):
            for col_idx in range(1, len(row_data) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = REVIEW_FILL

        # Colorear celda de respuesta correcta en verde
        if excel_format != ExcelFormatEnum.EXAM:
            correct_col = headers.index("Respuesta Correcta") + 1
            ws.cell(row=row_idx, column=correct_col).fill = CORRECT_FILL

    wb.save(excel_path)
    return excel_path


def _build_row(
    num: int,
    q: Question,
    correct_text: str,
    excel_format: ExcelFormatEnum
) -> list:
    """Construye los valores de una fila segun el formato."""
    base = [
        num,
        q.question,
    ]
    options = [
        _strip_option_prefix(q.answer1),
        _strip_option_prefix(q.answer2),
        _strip_option_prefix(q.answer3),
        _strip_option_prefix(q.answer4) if q.answer4 else "",
    ]

    if excel_format == ExcelFormatEnum.EXAM:
        return base + options

    if excel_format == ExcelFormatEnum.WITH_SOLUTIONS:
        return base + options + [
            correct_text,
            q.tip or "",
            q.article or "",
        ]

    # STUDY_GUIDE
    return base + options + [
        correct_text,
        q.tip or "",
        q.article or "",
        getattr(q, "difficulty", "") or "",
        getattr(q, "difficulty_reason", "") or "",
        "SI" if getattr(q, "needs_manual_review", False) else "NO",
        getattr(q, "review_comment", "") or "",
        q.llm_model or "",
        q.faithfulness_score if q.faithfulness_score is not None else "",
        q.relevancy_score if q.relevancy_score is not None else "",
        round(q.generation_time, 2) if q.generation_time else "",
        round(getattr(q, "review_time", None) or 0, 2) if getattr(q, "review_time", None) else "",
    ]


# ==========================================
# 6. CONSTRUCCION DEL GRAFO
# ==========================================

def build_agent_f():
    """Construye y compila el grafo del Agente F."""
    workflow = StateGraph(AgentFState)
    workflow.add_node("group_by_topic", group_by_topic_node)
    workflow.add_node("generate_excels", generate_excels_node)

    workflow.set_entry_point("group_by_topic")
    workflow.add_edge("group_by_topic", "generate_excels")
    workflow.add_edge("generate_excels", END)

    return workflow.compile()


# ==========================================
# 7. CLASE PRINCIPAL DEL AGENTE
# ==========================================

class ExcelGeneratorAgent:
    """Agente F: Genera Excels de examenes por tema."""

    def __init__(self):
        self.graph = build_agent_f()
        logger.info("Initialized ExcelGeneratorAgent")

    def generate_excels(
        self,
        questions: List[Question],
        output_dir: Path = None,
        excel_format: ExcelFormatEnum = ExcelFormatEnum.WITH_SOLUTIONS,
        topic_filter: Optional[int] = None,
    ) -> List[ExcelMetadata]:
        """Genera Excels agrupados por tema.

        Args:
            questions: Lista de preguntas
            output_dir: Directorio de salida
            excel_format: Formato del Excel
            topic_filter: Filtrar solo un tema (opcional)

        Returns:
            Lista de metadata de Excels generados
        """
        if output_dir is None:
            output_dir = Path("output/excels")

        log_separator("AGENTE F: EXCEL GENERATOR", "=")

        initial_state = {
            "questions": questions,
            "output_dir": output_dir,
            "excel_format": excel_format,
            "topic_filter": topic_filter,
            "questions_by_topic": {},
            "excel_metadata": [],
            "decisions_log": [],
            "node_count": 0
        }

        final_state = self.graph.invoke(initial_state)

        print("\nRecorrido del grafo:")
        for log in final_state.get("decisions_log", []):
            print(f"   {log['timestamp'][-8:]} | {log['node']:25} | {log['decision']}")

        return final_state.get("excel_metadata", [])


# ==========================================
# 8. FUNCION DE CONVENIENCIA
# ==========================================

def generate_excels_from_questions(
    questions: List[Question],
    output_dir: str = "output/excels",
    excel_format: str = "with_solutions",
    topic_filter: Optional[int] = None,
) -> List[ExcelMetadata]:
    """Funcion standalone para generar Excels."""
    agent = ExcelGeneratorAgent()
    format_enum = ExcelFormatEnum(excel_format)
    return agent.generate_excels(
        questions=questions,
        output_dir=Path(output_dir),
        excel_format=format_enum,
        topic_filter=topic_filter,
    )


__all__ = [
    'ExcelGeneratorAgent',
    'ExcelFormatEnum',
    'generate_excels_from_questions',
]